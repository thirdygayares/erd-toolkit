from __future__ import annotations

from collections import defaultdict
import csv
from datetime import datetime, timezone
import io
import re

from psycopg.types.json import Jsonb

from app.core.context import RequestContext
from app.core.db import Database
from app.core.errors import ValidationError
from app.features.export import sql
from app.features.export.schemas import ExportDictionaryRequest, ExportSqlRequest


class ExportService:
    _IDENTITY_DEFAULT_RE = re.compile(
        r"^generated\s+(always|by\s+default)\s+as\s+identity(?:\s*\(.*\))?$",
        re.IGNORECASE,
    )
    _IDENTITY_COMPATIBLE_TYPES = {
        "smallint",
        "int2",
        "integer",
        "int",
        "int4",
        "bigint",
        "int8",
    }
    _SAFE_IDENTIFIER_RE = re.compile(r"^[a-z_][a-z0-9_]*$")
    _RESERVED_WORDS = {
        "all",
        "alter",
        "and",
        "any",
        "as",
        "asc",
        "between",
        "by",
        "case",
        "check",
        "constraint",
        "create",
        "default",
        "delete",
        "desc",
        "distinct",
        "drop",
        "exists",
        "false",
        "foreign",
        "from",
        "group",
        "having",
        "in",
        "index",
        "insert",
        "into",
        "is",
        "join",
        "key",
        "limit",
        "not",
        "null",
        "on",
        "or",
        "order",
        "primary",
        "references",
        "select",
        "set",
        "table",
        "true",
        "union",
        "unique",
        "update",
        "user",
        "using",
        "values",
        "where",
    }

    def __init__(self, db: Database) -> None:
        self.db = db

    def export_sql(self, diagram_id: str, payload: ExportSqlRequest, ctx: RequestContext) -> dict:
        with self.db.connection() as conn:
            self.db.apply_request_context(conn, ctx)
            with conn.cursor() as cur:
                cur.execute(sql.CREATE_EXPORT_JOB, {"diagram_id": diagram_id})
                export_job_id = cur.fetchone()["export_job_id"]

            try:
                sql_output, statement_count = self._generate_sql(
                    conn,
                    diagram_id,
                    payload.target_schema,
                    payload.source_schema_names,
                    payload.export_all_schemas,
                )
                with conn.cursor() as cur:
                    cur.execute(
                        sql.MARK_EXPORT_SUCCESS,
                        {
                            "export_job_id": export_job_id,
                            "sql_output": sql_output,
                            "diff_summary": Jsonb({"statement_count": statement_count}),
                        },
                    )
                return {
                    "export_job_id": export_job_id,
                    "status": "success",
                    "statement_count": statement_count,
                    "sql_output": sql_output,
                }
            except Exception as exc:
                with conn.cursor() as cur:
                    cur.execute(
                        sql.MARK_EXPORT_FAILED,
                        {
                            "export_job_id": export_job_id,
                            "error_text": str(exc),
                        },
                    )
                raise

    def export_dictionary(
        self,
        diagram_id: str,
        payload: ExportDictionaryRequest,
        ctx: RequestContext,
    ) -> dict:
        with self.db.connection() as conn:
            self.db.apply_request_context(conn, ctx)
            tables, relationships, custom_types, columns_by_table = self._load_export_entities(
                conn,
                diagram_id,
                source_schema_names=payload.source_schema_names,
                export_all_schemas=payload.export_all_schemas,
            )

        rows = self._build_dictionary_rows(
            tables=tables,
            relationships=relationships,
            custom_types=custom_types,
            columns_by_table=columns_by_table,
            layout=payload.layout,
            include_enums=payload.include_enums,
        )

        if payload.file_type == "csv":
            content = self._render_dictionary_csv_bytes(rows)
            content_type = "text/csv; charset=utf-8"
        else:
            content = self._render_dictionary_xlsx_bytes(rows, payload.layout)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H-%M-%SZ")
        filename = f"erd_data_dictionary_{timestamp}.{payload.file_type}"

        return {
            "filename": filename,
            "content": content,
            "content_type": content_type,
        }

    def _generate_sql(
        self,
        conn,
        diagram_id: str,
        target_schema: str,
        source_schema_names: list[str] | None = None,
        export_all_schemas: bool = True,
    ) -> tuple[str, int]:
        tables, relationships, custom_types, columns_by_table = self._load_export_entities(
            conn,
            diagram_id,
            source_schema_names=source_schema_names or [],
            export_all_schemas=export_all_schemas,
        )

        column_lookup: dict[str, str] = {}
        column_meta_lookup: dict[str, dict] = {}
        export_table_lookup = self._build_export_table_lookup(tables)

        for table in tables:
            cols = columns_by_table.get(str(table["table_id"]), [])
            for col in cols:
                column_id = str(col["column_id"])
                rendered_type = self._render_type_sql(col)
                column_lookup[column_id] = col["column_name"]
                column_meta_lookup[column_id] = {
                    "type_sql": rendered_type,
                    "is_primary_key": bool(col.get("is_primary_key")),
                    "is_unique": bool(col.get("is_unique")),
                }

        statements: list[str] = []
        warnings: list[str] = []

        for custom_type in custom_types:
            enum_values = custom_type.get("enum_values") or []
            if not enum_values:
                warnings.append(
                    "-- WARNING: skipped enum type "
                    f"{self._q(custom_type['type_name'])} because it has no values."
                )
                continue
            statements.append(
                self._render_custom_type_sql(
                    target_schema=target_schema,
                    type_name=custom_type["type_name"],
                    enum_values=enum_values,
                )
            )

        for table in tables:
            export_table_name = export_table_lookup.get(str(table["table_id"]), table["table_name"])
            cols = columns_by_table.get(str(table["table_id"]), [])
            column_defs: list[str] = []
            pk_columns: list[str] = []
            for col in cols:
                type_sql = self._render_type_sql(col)
                pieces = [f"{self._q(col['column_name'])} {type_sql}"]
                default_clause, warning = self._render_default_clause(
                    default_sql=col["default_sql"],
                    type_sql=type_sql,
                    table_name=export_table_name,
                    column_name=col["column_name"],
                )
                if default_clause:
                    pieces.append(default_clause)
                if warning:
                    warnings.append(warning)
                if not col["is_nullable"]:
                    pieces.append("NOT NULL")
                column_defs.append(" ".join(pieces))
                if col["is_primary_key"]:
                    pk_columns.append(self._q(col["column_name"]))
                if col.get("is_unique") and not col.get("is_primary_key"):
                    column_defs.append(f"UNIQUE ({self._q(col['column_name'])})")

            if pk_columns:
                column_defs.append(f"PRIMARY KEY ({', '.join(pk_columns)})")

            if column_defs:
                create_table_sql = (
                    f"CREATE TABLE IF NOT EXISTS {self._q(target_schema)}.{self._q(export_table_name)} (\n"
                    + "  "
                    + ",\n  ".join(column_defs)
                    + "\n);"
                )
            else:
                create_table_sql = (
                    f"CREATE TABLE IF NOT EXISTS {self._q(target_schema)}.{self._q(export_table_name)} ();"
                )
            statements.append(create_table_sql)

        for rel in relationships:
            from_table = export_table_lookup.get(str(rel["from_table_id"]))
            to_table = export_table_lookup.get(str(rel["to_table_id"]))
            from_column = column_lookup.get(str(rel["from_column_id"]))
            to_column = column_lookup.get(str(rel["to_column_id"]))
            if not from_table or not to_table or not from_column or not to_column:
                continue
            from_meta = column_meta_lookup.get(str(rel["from_column_id"]))
            to_meta = column_meta_lookup.get(str(rel["to_column_id"]))
            if not from_meta or not to_meta:
                continue

            if not (to_meta["is_primary_key"] or to_meta["is_unique"]):
                warnings.append(
                    "-- WARNING: skipped foreign key "
                    f"{self._q(rel['name'])} because referenced column "
                    f"{self._q(to_table)}.{self._q(to_column)} is not PRIMARY KEY or UNIQUE."
                )
                continue

            if not self._types_are_fk_compatible(from_meta["type_sql"], to_meta["type_sql"]):
                warnings.append(
                    "-- WARNING: skipped foreign key "
                    f"{self._q(rel['name'])} due to incompatible types "
                    f"{from_meta['type_sql']} -> {to_meta['type_sql']}."
                )
                continue

            statements.append(
                (
                    f"ALTER TABLE {self._q(target_schema)}.{self._q(from_table)} "
                    f"ADD CONSTRAINT {self._q(rel['name'])} "
                    f"FOREIGN KEY ({self._q(from_column)}) "
                    f"REFERENCES {self._q(target_schema)}.{self._q(to_table)} ({self._q(to_column)}) "
                    f"ON UPDATE {rel['on_update_action']} ON DELETE {rel['on_delete_action']};"
                )
            )

        output_parts: list[str] = []
        if warnings:
            output_parts.append("\n".join(warnings))
        if statements:
            output_parts.append("\n\n".join(statements))

        final_sql = "\n\n".join(output_parts) + ("\n" if output_parts else "")
        return final_sql, len(statements)

    def _load_export_entities(
        self,
        conn,
        diagram_id: str,
        *,
        source_schema_names: list[str],
        export_all_schemas: bool,
    ) -> tuple[list[dict], list[dict], list[dict], dict[str, list[dict]]]:
        with conn.cursor() as cur:
            cur.execute(sql.GET_TABLES, {"diagram_id": diagram_id})
            tables = cur.fetchall()

            selected_schema_names = self._resolve_source_schema_names(
                tables,
                source_schema_names=source_schema_names,
                export_all_schemas=export_all_schemas,
            )

            tables = [
                table for table in tables if table["schema_name"] in selected_schema_names
            ]
            tables.sort(key=lambda table: (table["schema_name"], table["table_name"]))

            cur.execute(sql.GET_CUSTOM_TYPES, {"diagram_id": diagram_id})
            custom_types = [
                custom_type
                for custom_type in cur.fetchall()
                if custom_type["schema_name"] in selected_schema_names
            ]
            custom_types.sort(
                key=lambda custom_type: (custom_type["schema_name"], custom_type["type_name"])
            )

            cur.execute(sql.GET_RELATIONSHIPS, {"diagram_id": diagram_id})
            relationships = cur.fetchall()
            selected_table_ids = {str(table["table_id"]) for table in tables}
            relationships = [
                rel
                for rel in relationships
                if str(rel["from_table_id"]) in selected_table_ids
                and str(rel["to_table_id"]) in selected_table_ids
            ]

            columns_by_table: dict[str, list[dict]] = defaultdict(list)
            for table in tables:
                cur.execute(sql.GET_COLUMNS, {"table_id": table["table_id"]})
                cols = cur.fetchall()
                cols.sort(key=lambda column: column.get("ordinal_position") or 0)
                columns_by_table[str(table["table_id"])] = cols

        return tables, relationships, custom_types, columns_by_table

    def _build_dictionary_rows(
        self,
        *,
        tables: list[dict],
        relationships: list[dict],
        custom_types: list[dict],
        columns_by_table: dict[str, list[dict]],
        layout: str,
        include_enums: bool,
    ) -> list[list[str]]:
        fk_lookup = self._build_fk_reference_lookup(tables, relationships, columns_by_table)
        if layout == "section_sheet":
            return self._build_section_dictionary_rows(
                tables=tables,
                columns_by_table=columns_by_table,
                fk_lookup=fk_lookup,
                custom_types=custom_types,
                include_enums=include_enums,
            )
        return self._build_flat_dictionary_rows(
            tables=tables,
            columns_by_table=columns_by_table,
            fk_lookup=fk_lookup,
            custom_types=custom_types,
            include_enums=include_enums,
        )

    def _build_flat_dictionary_rows(
        self,
        *,
        tables: list[dict],
        columns_by_table: dict[str, list[dict]],
        fk_lookup: dict[str, list[str]],
        custom_types: list[dict],
        include_enums: bool,
    ) -> list[list[str]]:
        rows: list[list[str]] = [
            [
                "Schema",
                "Table",
                "Field",
                "Type",
                "Not Null",
                "Default",
                "Example",
                "Unique",
                "PK",
                "FK",
                "Description",
            ]
        ]

        for table in tables:
            table_id = str(table["table_id"])
            for column in columns_by_table.get(table_id, []):
                column_id = str(column["column_id"])
                fk_references = fk_lookup.get(f"{table_id}:{column_id}", [])
                rows.append(
                    [
                        table["schema_name"],
                        table["table_name"],
                        column["column_name"],
                        self._render_type_sql(column),
                        self._to_yes_no(not column.get("is_nullable", True)),
                        str(column.get("default_sql") or ""),
                        str(column.get("example_value") or ""),
                        self._to_yes_no(bool(column.get("is_unique"))),
                        self._to_yes_no(bool(column.get("is_primary_key"))),
                        " | ".join(fk_references),
                        str(column.get("comment_text") or ""),
                    ]
                )

        if include_enums and custom_types:
            rows.append([])
            rows.append(["Enums"])
            rows.append(["Schema", "Type", "Value"])
            for custom_type in custom_types:
                enum_values = custom_type.get("enum_values") or []
                if not enum_values:
                    rows.append([custom_type["schema_name"], custom_type["type_name"], ""])
                    continue
                for enum_value in enum_values:
                    rows.append([custom_type["schema_name"], custom_type["type_name"], str(enum_value)])

        return rows

    def _build_section_dictionary_rows(
        self,
        *,
        tables: list[dict],
        columns_by_table: dict[str, list[dict]],
        fk_lookup: dict[str, list[str]],
        custom_types: list[dict],
        include_enums: bool,
    ) -> list[list[str]]:
        rows: list[list[str]] = []

        for table in tables:
            table_id = str(table["table_id"])
            rows.append([f"{table['schema_name']}.{table['table_name']}", "", "", "", "", "", "", ""])
            rows.append(["Key", "Field", "Type", "Not Null", "Default", "Description", "Example", "FK"])

            for column in columns_by_table.get(table_id, []):
                column_id = str(column["column_id"])
                fk_references = fk_lookup.get(f"{table_id}:{column_id}", [])
                rows.append(
                    [
                        self._build_key_label(
                            is_primary_key=bool(column.get("is_primary_key")),
                            has_foreign_key=bool(fk_references),
                        ),
                        column["column_name"],
                        self._render_type_sql(column),
                        "NOT NULL" if not column.get("is_nullable", True) else "NULLABLE",
                        str(column.get("default_sql") or ""),
                        str(column.get("comment_text") or ""),
                        str(column.get("example_value") or ""),
                        " | ".join(fk_references),
                    ]
                )

            rows.append([])

        if include_enums and custom_types:
            rows.append(["ENUMS", "", "", "", "", "", "", ""])
            rows.append(["Schema", "Type", "Value", "", "", "", "", ""])
            for custom_type in custom_types:
                enum_values = custom_type.get("enum_values") or []
                if not enum_values:
                    rows.append([custom_type["schema_name"], custom_type["type_name"], "", "", "", "", "", ""])
                    continue
                for enum_value in enum_values:
                    rows.append(
                        [
                            custom_type["schema_name"],
                            custom_type["type_name"],
                            str(enum_value),
                            "",
                            "",
                            "",
                            "",
                            "",
                        ]
                    )

        return rows

    @staticmethod
    def _build_fk_reference_lookup(
        tables: list[dict],
        relationships: list[dict],
        columns_by_table: dict[str, list[dict]],
    ) -> dict[str, list[str]]:
        table_lookup = {str(table["table_id"]): table for table in tables}
        column_name_lookup: dict[str, dict[str, str]] = {}
        for table_id, columns in columns_by_table.items():
            column_name_lookup[table_id] = {
                str(column["column_id"]): str(column["column_name"])
                for column in columns
            }

        fk_lookup: dict[str, list[str]] = defaultdict(list)
        for relationship in relationships:
            from_table_id = str(relationship["from_table_id"])
            from_column_id = str(relationship["from_column_id"])
            to_table_id = str(relationship["to_table_id"])
            to_column_id = str(relationship["to_column_id"])

            target_table = table_lookup.get(to_table_id)
            target_column_name = column_name_lookup.get(to_table_id, {}).get(to_column_id)
            if not target_table or not target_column_name:
                continue

            reference = (
                f"{target_table['schema_name']}.{target_table['table_name']}.{target_column_name}"
            )
            key = f"{from_table_id}:{from_column_id}"
            fk_lookup[key].append(reference)

        normalized_lookup: dict[str, list[str]] = {}
        for key, references in fk_lookup.items():
            normalized_lookup[key] = sorted(set(references))
        return normalized_lookup

    @staticmethod
    def _to_yes_no(value: bool) -> str:
        return "Yes" if value else "No"

    @staticmethod
    def _build_key_label(*, is_primary_key: bool, has_foreign_key: bool) -> str:
        if is_primary_key and has_foreign_key:
            return "PK, FK"
        if is_primary_key:
            return "PK"
        if has_foreign_key:
            return "FK"
        return ""

    @staticmethod
    def _render_dictionary_csv_bytes(rows: list[list[str]]) -> bytes:
        buffer = io.StringIO(newline="")
        writer = csv.writer(buffer)
        writer.writerows(rows)
        return buffer.getvalue().encode("utf-8-sig")

    def _render_dictionary_xlsx_bytes(self, rows: list[list[str]], layout: str) -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise ValidationError(
                "XLSX export requires openpyxl. Install backend dependencies first."
            ) from exc

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Dictionary"

        for row_index, row in enumerate(rows, start=1):
            for column_index, value in enumerate(row, start=1):
                worksheet.cell(row=row_index, column=column_index, value=value)

        max_col = max((len(row) for row in rows), default=1)
        gray_fill = PatternFill(fill_type="solid", fgColor="FFD9D9D9")
        green_fill = PatternFill(fill_type="solid", fgColor="FF00FF00")
        bold_font = Font(bold=True)

        if layout == "table_grid":
            for column_index in range(1, max_col + 1):
                cell = worksheet.cell(row=1, column=column_index)
                cell.fill = gray_fill
                cell.font = bold_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
            worksheet.freeze_panes = "A2"
        else:
            for row_index, row in enumerate(rows, start=1):
                non_empty_values = [value for value in row if value not in (None, "")]
                if not non_empty_values:
                    continue

                first_value = str(row[0] if row else "").strip()
                is_section_title = len(non_empty_values) == 1 and (
                    "." in first_value or first_value.upper() == "ENUMS"
                )
                is_header_row = first_value.upper() in {"KEY", "SCHEMA"} and len(non_empty_values) >= 3

                if is_section_title:
                    for column_index in range(1, max_col + 1):
                        cell = worksheet.cell(row=row_index, column=column_index)
                        cell.fill = green_fill
                        cell.font = bold_font
                elif is_header_row:
                    for column_index in range(1, max_col + 1):
                        cell = worksheet.cell(row=row_index, column=column_index)
                        cell.fill = gray_fill
                        cell.font = bold_font

        for column_index in range(1, max_col + 1):
            max_length = 0
            for row_index in range(1, worksheet.max_row + 1):
                value = worksheet.cell(row=row_index, column=column_index).value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
            worksheet.column_dimensions[worksheet.cell(row=1, column=column_index).column_letter].width = max(
                10,
                min(80, max_length + 2),
            )

        stream = io.BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    @staticmethod
    def _q(identifier: str) -> str:
        if (
            ExportService._SAFE_IDENTIFIER_RE.match(identifier)
            and identifier not in ExportService._RESERVED_WORDS
        ):
            return identifier
        return '"' + identifier.replace('"', '""') + '"'

    @staticmethod
    def _normalize_export_identifier(value: str | None) -> str | None:
        if not value:
            return None
        normalized = re.sub(r"[^a-z0-9_]+", "_", value.strip().lower()).strip("_")
        if not normalized or len(normalized) > 63:
            return None
        return normalized

    def _build_export_table_lookup(self, tables: list[dict]) -> dict[str, str]:
        preferred_names: dict[str, str] = {}
        name_usage: dict[str, int] = defaultdict(int)

        for table in tables:
            table_id = str(table["table_id"])
            candidate = self._normalize_export_identifier(table.get("display_name"))
            if candidate:
                preferred_names[table_id] = candidate
                name_usage[candidate] += 1

        export_names: dict[str, str] = {}
        for table in tables:
            table_id = str(table["table_id"])
            preferred = preferred_names.get(table_id)
            if preferred and name_usage.get(preferred, 0) == 1:
                export_names[table_id] = preferred
            else:
                export_names[table_id] = table["table_name"]

        return export_names

    @staticmethod
    def _render_type_sql(col: dict) -> str:
        data_type = str(col["data_type"]).strip()
        is_user_defined = data_type.replace("[]", "") == "USER-DEFINED"
        type_sql = (
            f"{col['udt_name']}{'[]' if data_type.endswith('[]') else ''}"
            if is_user_defined and col["udt_name"]
            else data_type
        )
        normalized = str(type_sql).strip()
        lower = normalized.lower()
        if lower == "varchar(n)":
            return "varchar"
        if normalized.endswith("?"):
            return normalized[:-1].strip()
        return normalized

    @staticmethod
    def _quote_literal(value: str) -> str:
        return "'" + value.replace("'", "''") + "'"

    def _render_custom_type_sql(
        self,
        *,
        target_schema: str,
        type_name: str,
        enum_values: list[str],
    ) -> str:
        labels = ", ".join(self._quote_literal(value) for value in enum_values)
        return (
            "DO $$\n"
            "BEGIN\n"
            f"  CREATE TYPE {self._q(target_schema)}.{self._q(type_name)} AS ENUM ({labels});\n"
            "EXCEPTION\n"
            "  WHEN duplicate_object THEN NULL;\n"
            "END $$;"
        )

    @classmethod
    def _is_identity_default(cls, default_sql: str) -> bool:
        return bool(cls._IDENTITY_DEFAULT_RE.match(default_sql.strip()))

    @classmethod
    def _is_identity_compatible_type(cls, type_sql: str) -> bool:
        canonical = cls._canonical_type(type_sql)
        if canonical.endswith("[]"):
            return False
        return canonical in cls._IDENTITY_COMPATIBLE_TYPES

    def _render_default_clause(
        self,
        *,
        default_sql: str | None,
        type_sql: str,
        table_name: str,
        column_name: str,
    ) -> tuple[str | None, str | None]:
        if not default_sql:
            return None, None

        normalized_default = " ".join(default_sql.strip().split())
        if not normalized_default:
            return None, None

        if self._is_identity_default(normalized_default):
            if self._is_identity_compatible_type(type_sql):
                if normalized_default.lower().startswith("generated always"):
                    return "GENERATED ALWAYS AS IDENTITY", None
                return "GENERATED BY DEFAULT AS IDENTITY", None

            warning = (
                f"-- WARNING: dropped identity clause for "
                f"{self._q(table_name)}.{self._q(column_name)} "
                f"because type {type_sql} is not integer-compatible."
            )
            return None, warning

        return f"DEFAULT {default_sql.strip()}", None

    @staticmethod
    def _canonical_type(type_sql: str) -> str:
        normalized = type_sql.strip().lower()
        if normalized.endswith("[]"):
            base = re.sub(r"\(.*\)", "", normalized[:-2]).strip()
            return f"{base}[]"
        base = re.sub(r"\(.*\)", "", normalized).strip()
        aliases = {
            "int": "integer",
            "int4": "integer",
            "integer": "integer",
            "bigint": "bigint",
            "int8": "bigint",
            "smallint": "smallint",
            "int2": "smallint",
            "bool": "boolean",
            "character varying": "varchar",
            "varchar": "varchar",
            "timestamp with time zone": "timestamptz",
            "timestamp without time zone": "timestamp",
        }
        return aliases.get(base, base)

    @classmethod
    def _types_are_fk_compatible(cls, from_type: str, to_type: str) -> bool:
        left = cls._canonical_type(from_type)
        right = cls._canonical_type(to_type)
        if left == right:
            return True
        return {left, right} == {"text", "varchar"}

    def _resolve_source_schema_names(
        self,
        tables: list[dict],
        *,
        source_schema_names: list[str],
        export_all_schemas: bool,
    ) -> set[str]:
        available = {table["schema_name"] for table in tables}
        if export_all_schemas:
            return available

        selected = {schema.strip() for schema in source_schema_names if schema and schema.strip()}
        if not selected:
            raise ValidationError(
                "No source schemas selected for export. Choose at least one schema or enable export-all."
            )

        unknown = sorted(selected - available)
        if unknown:
            raise ValidationError(
                f"Unknown source schema selection: {', '.join(unknown)}"
            )
        return selected
